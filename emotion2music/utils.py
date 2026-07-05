"""
Utility Functions Module.
Contains helper functions for parsing LLLL (List of Lists of Lists of Lists) strings,
calculating musical durations, extracting metadata, and exporting logs.
"""

import re
import csv
import json
from openpyxl import Workbook, load_workbook


def prepare_osc_list(llll_string):
    """
    Formats an LLLL string into a flat list of elements suitable for OSC transmission.
    Properly spaces brackets and converts numeric strings to integers where applicable.
    """
    if not llll_string:
        return []

    if isinstance(llll_string, list):
        llll_string = str(llll_string).replace("'", "").replace('"', "").replace(",", "")

    spaced_text = llll_string.replace('[', ' [ ').replace(']', ' ] ')

    return [int(e) if e.replace('-', '').isdigit() else e for e in spaced_text.split()]


def extract_llll_blocks(text):
    """
    Parses a formatted LLLL string and extracts the top-level list blocks.
    """
    text = text.strip()
    blocks, depth, start = [], 0, 0

    for i, char in enumerate(text):
        if char == '[':
            if depth == 0:
                start = i
            depth += 1
        elif char == ']':
            depth -= 1
            if depth == 0:
                blocks.append(text[start:i + 1])

    return blocks


def concatenate_lists(existing_llll, new_llll):
    """
    Concatenates two LLLL strings by merging their internal contents voice by voice.
    This is used to append newly generated measures to the global score.
    """
    if not existing_llll:
        return new_llll

    blocks_old = extract_llll_blocks(existing_llll)
    blocks_new = extract_llll_blocks(new_llll)
    result = []

    for i in range(max(len(blocks_old), len(blocks_new))):
        old_inner = blocks_old[i].strip()[1:-1].strip() if i < len(blocks_old) else ""
        new_inner = blocks_new[i].strip()[1:-1].strip() if i < len(blocks_new) else ""
        result.append(f"[ {old_inner} {new_inner} ]")

    return " ".join(result)


def calculate_total_duration(measures_llll):
    """
    Calculates the absolute playback duration (in seconds) of a musical segment
    based on its time signature and BPM.
    """
    if not measures_llll:
        return 0.0

    tokens = str(measures_llll).replace('[', ' ').replace(']', ' ').split()
    nums = [float(t) for t in tokens if t.replace('.', '').isdigit()]

    duration = 0.0
    step = 3  # [numerator, denominator, bpm]

    for i in range(0, len(nums), step):
        if i + 2 < len(nums):
            beats_per_measure = nums[i] * 4.0 / nums[i + 1]
            seconds_per_beat = 60.0 / nums[i + 2]
            duration += beats_per_measure * seconds_per_beat

    return duration


def extract_bpm(measures_llll):
    """
    Extracts the first valid BPM value from a measures array.
    """
    if not measures_llll:
        return None

    tokens = str(measures_llll).replace('[', ' ').replace(']', ' ').split()
    nums = [float(t) for t in tokens if t.replace('.', '').isdigit()]

    if len(nums) >= 3:
        return nums[2]

    return None


def transpose_pitch_list(pitch_llll, octave_transposition=2):
    """
    Transposes all pitches in an LLLL string by a specified number of octaves.
    Ensures that the resulting octaves remain within standard MIDI bounds (-1 to 9).
    """
    if not pitch_llll:
        return ""

    def shift_octave(match):
        raw_note = match.group(1)
        original_octave = int(match.group(2))
        new_octave = max(-1, min(9, original_octave + octave_transposition))
        return f"{raw_note}{new_octave}"

    return re.sub(r'([A-G][b#]?)(\d+)', shift_octave, str(pitch_llll))


def export_logs_and_metrics(musical_data, metric_data, csv_file, xlsx_file, diary_file, json_file, logger):
    """
    Handles the physical saving of generated metrics and data to CSV, Excel,
    and text (diary/JSON) files.
    """
    current_time_str = metric_data.get("Timestamp", "Unknown Time")
    emotion = metric_data.get("Emotion", "Unknown Emotion")

    # CSV Export
    try:
        file_exists = csv_file.exists()
        with open(csv_file, 'a', newline='', encoding='utf-8') as f_csv:
            writer = csv.DictWriter(f_csv, fieldnames=metric_data.keys(), delimiter=';')
            if not file_exists:
                writer.writeheader()
            writer.writerow(metric_data)
    except Exception as e:
        logger.error(f"CSV writing error: {e}")

    # Excel Export
    try:
        if not xlsx_file.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(list(metric_data.keys()))
        else:
            wb = load_workbook(xlsx_file)
            ws = wb.active

        ws.append(list(metric_data.values()))
        wb.save(xlsx_file)
    except Exception as e:
        logger.error(f"Excel writing error: {e}")

    # JSON and Diary Export
    if musical_data:
        try:
            if "composer_diary" in musical_data:
                with open(diary_file, 'a', encoding='utf-8') as f_diary:
                    f_diary.write(f"[{current_time_str} | Emotion: {emotion}]\n{musical_data['composer_diary']}\n\n")

            with open(json_file, 'a', encoding='utf-8') as f_json:
                f_json.write(
                    f"// --- Generated JSON: {current_time_str} | {emotion} ---\n{json.dumps(musical_data, indent=2)}\n\n")
        except Exception as e:
            logger.error(f"Text file writing error: {e}")