"""
Evaluation Table Generator.
Parses experiment logs, JSON evaluations, and diary texts to generate a
consolidated, formatted Excel benchmark report.
"""

import pandas as pd
import json
import os
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


def parse_experiment_log(log_path):
    """Parses a single experiment log file to extract metadata and metric blocks."""
    blocks = []
    instruments = []
    customization_prompt = ""
    session_temp = ""

    if not log_path or not os.path.exists(log_path):
        return pd.DataFrame()

    with open(log_path, 'r', encoding='utf-8', errors='ignore') as f:
        log_lines = f.readlines()

    current_block = {}
    for line in log_lines:
        # Extract instrument voice assignments
        voice_match = re.search(r"Voice \d+ set to:\s*\d+\s*(.+)", line)
        if voice_match:
            instruments.append(voice_match.group(1).strip())

        # Extract instrument array lists
        instruments_match = re.search(r"Instruments\s*=\s*(\[.*?\])", line)
        if instruments_match:
            try:
                parsed_instr = json.loads(instruments_match.group(1).replace("'", '"'))
                if isinstance(parsed_instr, list):
                    instruments.extend(parsed_instr)
            except json.JSONDecodeError:
                pass

        # Extract textual instruction updates
        prompt_match = re.search(r"Updated text instructions:\s*(.+)", line)
        if prompt_match:
            customization_prompt = prompt_match.group(1).strip()

        # Extract session temperature
        temp_match = re.search(r"with temperature:\s*(\d+\.\d+)", line)
        if temp_match:
            session_temp = temp_match.group(1).strip()

        # Extract standardized metrics
        if "[THESIS-METRIC]" in line:
            parts = line.split("[THESIS-METRIC]")[1].split("|")
            for part in parts:
                if "=" in part:
                    key, value = part.split("=", 1)
                    current_block[key.strip()] = value.strip().strip("'\"")

            # Append the block when the task is marked as completed
            if "Task" in current_block and current_block["Task"] == "Completed":
                blocks.append(current_block)
                current_block = {}

    # Deduplicate instruments maintaining order
    seen = set()
    unique_instruments = [x for x in instruments if not (x in seen or seen.add(x))]

    base_df = pd.DataFrame(blocks)
    if not base_df.empty:
        base_df["Instruments_Used"] = ", ".join(unique_instruments) if unique_instruments else "Not detected"
        base_df["Genre_Personalization_Prompt"] = customization_prompt if customization_prompt else "Default"
        base_df["Session_Temperature"] = session_temp if session_temp else (
            base_df["Temperature"].iloc[0] if "Temperature" in base_df.columns else "N/A"
        )

    return base_df


def parse_diary(diary_path):
    """Extracts qualitative notes from the composer diary file."""
    if not diary_path or not os.path.exists(diary_path):
        return pd.DataFrame()

    with open(diary_path, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()

    # Regex matches the standardized diary format generated in utils.py
    pattern = re.compile(r'\[(.*?) \| Emotion: (.*?)\](.*?)(?=\n\[|$)', re.DOTALL)
    matches = pattern.findall(content)

    return pd.DataFrame([{"Musical_Diary": text.strip()} for _, _, text in matches])


def format_excel(excel_path):
    """Applies professional formatting, color scales, and structural merging to the Excel report."""
    wb = load_workbook(excel_path)

    # Apply formatting to all worksheets (Master + Individual Groups)
    for worksheet in wb.worksheets:
        worksheet.freeze_panes = "C2"  # Freeze view after Group and Experiment columns

        if worksheet.max_row > 1 and worksheet.max_column > 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill(start_color="2A3F54", end_color="2A3F54", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        border_thin = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        thick_bottom_line = Side(style='medium', color='2A3F54')

        # Style headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        text_columns = [
            'Musical_Diary',
            'Critical_Analysis',
            'Global_transitions_and_flow_analysis',
            'Global_advanced_qualitative_analysis'
        ]

        score_rule = ColorScaleRule(
            start_type='num', start_value=1, start_color='F8696B',
            mid_type='num', mid_value=3, mid_color='FFEB84',
            end_type='num', end_value=5, end_color='63BE7B'
        )

        columns_to_merge_per_experiment = []
        group_column = None

        # Adjust column widths and map columns for merging/formatting
        for col in worksheet.columns:
            col_name = str(col[0].value) if col[0].value else ""
            col_idx = col[0].column
            col_letter = get_column_letter(col_idx)

            if col_name == "Experiment_Group":
                group_column = col_idx
                worksheet.column_dimensions[col_letter].width = 25
            elif col_name == "Experiment_Name":
                columns_to_merge_per_experiment.append(col_idx)
                worksheet.column_dimensions[col_letter].width = 35
            elif col_name.startswith("Global_"):
                columns_to_merge_per_experiment.append(col_idx)
                worksheet.column_dimensions[col_letter].width = 60 if col_name in text_columns else 22

            if col_name in text_columns:
                worksheet.column_dimensions[col_letter].width = 60
            elif 'Score_' in col_name or 'score' in col_name.lower():
                worksheet.column_dimensions[col_letter].width = 15
                range_str = f"{col_letter}2:{col_letter}{worksheet.max_row}"
                worksheet.conditional_formatting.add(range_str, score_rule)
            elif col_name not in ["Experiment_Group", "Experiment_Name"] and not col_name.startswith("Global_"):
                worksheet.column_dimensions[col_letter].width = 22

            # Apply cell borders and alignment
            for cell in col[1:]:
                align_h = 'center' if ('Score_' in col_name or 'score' in col_name.lower()) else 'left'
                cell.alignment = Alignment(horizontal=align_h, vertical='top', wrap_text=True)
                cell.border = border_thin

        # --- 1. MERGE CELLS FOR "EXPERIMENT NAME" AND GLOBAL METRICS ---
        current_exp = None
        start_row_exp = 2
        max_row = worksheet.max_row
        exp_name_column = 2  # Default mapped to the second column

        for row in range(2, max_row + 2):
            exp_name = worksheet.cell(row=row, column=exp_name_column).value if row <= max_row else None

            if exp_name != current_exp:
                if current_exp is not None:
                    end_row = row - 1
                    if end_row >= start_row_exp:
                        for col_idx in columns_to_merge_per_experiment:
                            if end_row > start_row_exp:
                                worksheet.merge_cells(
                                    start_row=start_row_exp, start_column=col_idx,
                                    end_row=end_row, end_column=col_idx
                                )
                            merged_cell = worksheet.cell(row=start_row_exp, column=col_idx)
                            merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # Add a thick bottom border to distinctly separate the runs
                    for c_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=end_row, column=c_idx)
                        cell.border = Border(
                            left=cell.border.left, right=cell.border.right,
                            top=cell.border.top, bottom=thick_bottom_line
                        )
                current_exp = exp_name
                start_row_exp = row

        # --- 2. HIERARCHICAL MERGE FOR "EXPERIMENT GROUP" ---
        if group_column:
            current_group = None
            start_row_grp = 2
            for row in range(2, max_row + 2):
                grp_name = worksheet.cell(row=row, column=group_column).value if row <= max_row else None

                if grp_name != current_group:
                    if current_group is not None:
                        end_row = row - 1
                        if end_row > start_row_grp:
                            worksheet.merge_cells(
                                start_row=start_row_grp, start_column=group_column,
                                end_row=end_row, end_column=group_column
                            )
                        merged_cell = worksheet.cell(row=start_row_grp, column=group_column)
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    current_group = grp_name
                    start_row_grp = row

    wb.save(excel_path)


def find_file(root, files, extension_or_suffix, exclude=None):
    """Helper method to locate specific files within a directory traversal."""
    if exclude is None:
        exclude = []

    for f in files:
        if any(f.endswith(excl) for excl in exclude):
            continue
        if f.endswith(extension_or_suffix):
            return os.path.join(root, f)

    return None


def extract_single_experiment_data(log_path, diary_path, json_path, exp_group, exp_name):
    """Compiles all data sources for a single experiment into one aligned DataFrame."""
    df_base = parse_experiment_log(log_path)
    df_diary = parse_diary(diary_path)

    with open(json_path, "r", encoding="utf-8") as f:
        parsed_json = json.load(f)

    flat_blocks = []
    if "block_evaluations" in parsed_json:
        for block in parsed_json["block_evaluations"]:
            row = {
                "Judge_Block": block.get("block"),
                "Judge_Emotion": block.get("emotion"),
                "Critical_Analysis": block.get("critical_analysis"),
            }
            if "scores" in block and isinstance(block["scores"], dict):
                for k, v in block["scores"].items():
                    row[f"Score_{k}"] = v
            flat_blocks.append(row)

    df_blocks = pd.DataFrame(flat_blocks)

    dataframes_to_merge = [df for df in [df_base, df_diary, df_blocks] if not df.empty]
    if dataframes_to_merge:
        df_merged = pd.concat([df.reset_index(drop=True) for df in dataframes_to_merge], axis=1)
    else:
        df_merged = pd.DataFrame()

    if not df_merged.empty:
        df_merged.insert(0, 'Experiment_Name', exp_name)
        df_merged.insert(0, 'Experiment_Group', exp_group)

        if "global_evaluation" in parsed_json:
            for k, v in parsed_json["global_evaluation"].items():
                df_merged[f"Global_{k}"] = v

        # Organize columns based on English keys
        priority_columns = [
            "Experiment_Group", "Experiment_Name", "Genre_Personalization_Prompt", "Instruments_Used",
            "Session_Temperature", "Backend", "Model", "Emotion", "Judge_Emotion",
            "Temperature", "BPM", "Voices", "Musical_Diary"
        ]

        actual_order = [col for col in priority_columns if col in df_merged.columns]
        remaining_columns = [col for col in df_merged.columns if col not in actual_order and col != "Task"]
        df_merged = df_merged[actual_order + remaining_columns]

    return df_merged


def process_all_experiments(base_folder="."):
    """Traverses the experiments directory, processes data, and outputs the final Excel report."""
    base_path = Path(base_folder)
    all_blocks = []

    print("Initiating experiment search...")

    for root, dirs, files in os.walk(base_path):
        if "evaluation.json" in files:
            rel_path = os.path.relpath(root, base_path)
            parts = Path(rel_path).parts

            if len(parts) > 1:
                exp_group = parts[0]
                exp_name = " - ".join(parts[1:])
            else:
                exp_group = "General"
                exp_name = parts[0]

            json_path = os.path.join(root, "evaluation.json")
            log_path = find_file(root, files, ".txt", exclude=["_jsons.txt", "_diaries.txt", "_score.txt"])

            if not log_path:
                log_path = find_file(root, files, ".log")

            diary_path = find_file(root, files, "_diaries.txt")

            if log_path:
                print(f"   Extracting data from: [{exp_group}] {exp_name}")
                try:
                    df = extract_single_experiment_data(log_path, diary_path, json_path, exp_group, exp_name)
                    if not df.empty:
                        all_blocks.append(df)
                except Exception as e:
                    print(f"   Error processing {exp_name}: {e}")

    if all_blocks:
        print("\nSorting and saving the global master file and individual group sheets...")

        all_blocks.sort(key=lambda x: (x['Experiment_Group'].iloc[0], x['Experiment_Name'].iloc[0]))
        df_final = pd.concat(all_blocks, ignore_index=True)
        output_path = os.path.join(base_folder, "Global_Benchmark_Report.xlsx")

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # 1. Save master report containing all data
            df_final.to_excel(writer, sheet_name="Master Report", index=False)

            # 2. Create separated sheets for each Experiment Group
            groups = df_final['Experiment_Group'].unique()
            for group in groups:
                # Sanitize sheet name for Excel limitations (no special characters, max 31 length)
                safe_sheet_name = re.sub(r'[\\*?:/\[\]]', '', str(group))[:31]
                df_group = df_final[df_final['Experiment_Group'] == group]
                df_group.to_excel(writer, sheet_name=safe_sheet_name, index=False)

        format_excel(output_path)
        print(f"Process completed successfully. The report has been saved to: {output_path}")
    else:
        print("\nNo valid data found to generate the global report.")


if __name__ == "__main__":
    print("Starting the generation of the Global Excel Benchmark Report...")
    process_all_experiments()